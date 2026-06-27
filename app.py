"""FamilyFinance - Flask Web Application"""

import os
import sys
from flask import Flask, render_template, request, redirect, flash, url_for, jsonify
from datetime import datetime, timedelta
from collections import defaultdict

import config
from services.record_service import (
    load_data, save_data, find_record, create_purchase, create_redeem,
    get_grouped_records, get_redeems_for_purchase, get_purchase_records
)
from utils.validators import validate_purchase, validate_redeem
from utils.helpers import calculate_auto_profit, calculate_real_rate, days_between, format_amount

# Support PyInstaller single-file mode
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__,
            template_folder=os.path.join(base_path, 'templates'),
            static_folder=os.path.join(base_path, 'static'))
app.secret_key = config.SECRET_KEY


# Make helper available in all templates
@app.context_processor
def utility_processor():
    return dict(calculate_auto_profit=calculate_auto_profit)


# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------
def calculate_monthly_profits(records):
    monthly_profits = defaultdict(float)
    monthly_labels = []
    for record in records:
        if record.get('type') == 'redeem' and 'actual_profit' in record:
            redeem_date = datetime.strptime(record['redeem_date'], '%Y-%m-%d')
            month_key = redeem_date.strftime('%Y-%m')
            if month_key not in monthly_labels:
                monthly_labels.append(month_key)
            monthly_profits[month_key] += record['actual_profit']
    monthly_labels.sort()
    return monthly_labels, [monthly_profits[m] for m in monthly_labels]


def calculate_yearly_profits(records):
    yearly_profits = defaultdict(float)
    yearly_labels = []
    for record in records:
        if record.get('type') == 'redeem' and 'actual_profit' in record:
            redeem_date = datetime.strptime(record['redeem_date'], '%Y-%m-%d')
            year_key = redeem_date.strftime('%Y')
            if year_key not in yearly_labels:
                yearly_labels.append(year_key)
            yearly_profits[year_key] += record['actual_profit']
    yearly_labels.sort()
    return yearly_labels, [yearly_profits[y] for y in yearly_labels]


def calculate_product_distribution(records):
    product_amounts = defaultdict(float)
    product_names = []
    for record in records:
        name = record['product_name']
        if name not in product_amounts:
            product_names.append(name)
        if record['type'] == 'purchase':
            product_amounts[name] += record['amount']
        elif record['type'] == 'redeem':
            product_amounts[name] += record['redeem_amount']
    return product_names, [product_amounts[n] for n in product_names]


# ---------------------------------------------------------------------------
# Dashboard helpers
# ---------------------------------------------------------------------------
def get_dashboard_stats(records):
    """Compute dashboard summary metrics."""
    purchases = get_purchase_records(records)
    redeems = [r for r in records if r.get('type') == 'redeem']

    total_invested = sum(r['amount'] for r in purchases)
    total_redeemed = sum(r.get('redeem_amount', 0) for r in redeems)
    total_profit = sum(r.get('actual_profit', 0) for r in redeems)
    holding_principal = total_invested - total_redeemed

    # Average real rate
    rates = []
    for r in redeems:
        days = days_between(r['purchase_date'], r['redeem_date'])
        if days > 0:
            rate = calculate_real_rate(r['actual_profit'], r['redeem_amount'], days)
            rates.append(rate)
    avg_rate = sum(rates) / len(rates) if rates else 0

    return {
        'total_invested': round(total_invested, 2),
        'holding_principal': round(max(holding_principal, 0), 2),
        'total_profit': round(total_profit, 2),
        'avg_rate': round(avg_rate, 2),
    }


def get_upcoming_maturities(records, days=30):
    """Get purchases ending within N days that are still holding."""
    today = datetime.now().date()
    threshold = today + timedelta(days=days)
    purchases = get_purchase_records(records)
    redeems = [r for r in records if r.get('type') == 'redeem']

    result = []
    for p in purchases:
        end = datetime.strptime(p['end_date'], '%Y-%m-%d').date()
        if today <= end <= threshold:
            redeemed = sum(r['redeem_amount'] for r in redeems if r.get('purchase_record_id') == p['id'])
            remaining = p['amount'] - redeemed
            if remaining > 0:
                result.append({
                    'product_name': p['product_name'],
                    'bank_name': p.get('bank_name', ''),
                    'end_date': p['end_date'],
                    'amount': remaining,
                    'days_remaining': (end - today).days,
                    'purchase_date': p['purchase_date'],
                })
    return sorted(result, key=lambda x: x['purchase_date'], reverse=True)


def get_recent_redeems(records, limit=5):
    """Get most recent redeem records."""
    redeems = sorted(
        [r for r in records if r.get('type') == 'redeem'],
        key=lambda x: x['purchase_date'],
        reverse=True,
    )
    for r in redeems[:limit]:
        days = days_between(r['purchase_date'], r['redeem_date'])
        r['_real_rate'] = round(calculate_real_rate(
            r['actual_profit'], r['redeem_amount'], days), 2) if days > 0 else 0
    return redeems[:limit]


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    records = load_data()
    grouped = get_grouped_records(records)

    # Sort dashboard lists by purchase date (newest first)
    grouped.sort(key=lambda g: g['purchase']['purchase_date'], reverse=True)

    stats = get_dashboard_stats(records)
    upcoming = get_upcoming_maturities(records)
    recent = get_recent_redeems(records)
    monthly_labels, monthly_profits = calculate_monthly_profits(records)

    return render_template('index.html',
                           grouped=grouped,
                           stats=stats,
                           upcoming=upcoming,
                           recent_redeems=recent,
                           records=records,
                           monthly_labels=monthly_labels,
                           monthly_profits=monthly_profits)


@app.route('/records')
def records_list():
    records = load_data()
    grouped = get_grouped_records(records)

    # View mode: card (default) or table
    view_mode = request.args.get('mode', 'card')

    # Search/filter from query params
    search = request.args.get('search', '').strip().lower()
    status_filter = request.args.get('status', '')

    # Sort options
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    if search:
        grouped = [
            g for g in grouped
            if search in g['purchase']['product_name'].lower()
            or search in g['purchase'].get('bank_name', '').lower()
        ]
    if status_filter:
        grouped = [g for g in grouped if g['status'] == status_filter]

    # Sort grouped records
    reverse = (sort_order == 'desc')
    if sort_by == 'date':
        grouped.sort(key=lambda g: g['purchase']['purchase_date'], reverse=reverse)
    elif sort_by == 'amount':
        grouped.sort(key=lambda g: g['purchase']['amount'], reverse=reverse)
    elif sort_by == 'rate':
        grouped.sort(key=lambda g: g['purchase']['annual_rate'], reverse=reverse)
    elif sort_by == 'status':
        status_order = {'holding': 0, 'partial': 1, 'expired': 2, 'completed': 3}
        grouped.sort(key=lambda g: status_order.get(g['status'], 99), reverse=reverse)

    # Pre-compute real_rate for each redeem record for the template
    for g in grouped:
        for r in g['redeems']:
            days = days_between(r['purchase_date'], r['redeem_date'])
            r['_real_rate'] = round(calculate_real_rate(
                r['actual_profit'], r['redeem_amount'], days), 2) if days > 0 else 0

    return render_template('records/list.html',
                           grouped=grouped,
                           search=search,
                           status_filter=status_filter,
                           view_mode=view_mode,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           total=len(grouped),
                           calculate_auto_profit=calculate_auto_profit)


@app.route('/add', methods=['GET', 'POST'])
def add_record():
    records = load_data()
    if request.method == 'POST':
        try:
            record_type = request.form.get('record_type', '')
            if not record_type:
                flash('请选择记录类型', 'error')
                return redirect('/add')

            if record_type == 'purchase':
                # Validate required fields
                product_name = request.form.get('product_name', '').strip()
                amount_str = request.form.get('amount', '').strip()
                rate_str = request.form.get('annual_rate', '').strip()
                duration_str = request.form.get('duration', '').strip()
                purchase_date = request.form.get('purchase_date', '').strip()

                if not product_name:
                    flash('请输入产品名称', 'error')
                    return redirect('/add')
                if not amount_str:
                    flash('请输入投资金额', 'error')
                    return redirect('/add')
                if not rate_str:
                    flash('请输入年化利率', 'error')
                    return redirect('/add')
                if not duration_str:
                    flash('请输入投资期限', 'error')
                    return redirect('/add')
                if not purchase_date:
                    flash('请选择购买日期', 'error')
                    return redirect('/add')

                amount = float(amount_str)
                annual_rate = float(rate_str) / 100.0
                duration = int(duration_str)

                record = create_purchase({
                    'product_name': product_name,
                    'amount': amount,
                    'annual_rate': annual_rate,
                    'duration': duration,
                    'purchase_date': purchase_date,
                    'bank_name': request.form.get('bank_name', ''),
                })
            else:
                purchase_record_id = request.form.get('purchase_record_id', '').strip()
                redeem_amount_str = request.form.get('redeem_amount', '').strip()
                redeem_date = request.form.get('redeem_date', '').strip()

                if not purchase_record_id:
                    flash('请选择关联的买入记录', 'error')
                    return redirect('/add')
                if not redeem_amount_str:
                    flash('请输入赎回金额', 'error')
                    return redirect('/add')
                if not redeem_date:
                    flash('请选择赎回日期', 'error')
                    return redirect('/add')

                purchase = find_record(records, purchase_record_id)
                if not purchase:
                    flash('买入记录未找到', 'error')
                    return redirect('/add')

                redeem_amount = float(redeem_amount_str)
                if redeem_amount > purchase['amount']:
                    flash('赎回金额不能超过原始金额', 'error')
                    return redirect('/add')

                record = create_redeem(request.form, purchase)

            records.append(record)
            save_data(records)
            flash('理财记录添加成功！', 'success')
            return redirect('/')

        except ValueError as e:
            flash(f'输入格式错误: {str(e)}，请检查数字格式是否正确', 'error')
            return redirect('/add')
        except Exception as e:
            flash(f'添加记录失败: {str(e)}', 'error')
            return redirect('/add')

    # Filter out completed (fully redeemed) purchase records for the redeem dropdown
    grouped = get_grouped_records(records)
    completed_ids = {g['purchase']['id'] for g in grouped if g['status'] == 'completed'}
    purchase_records = [r for r in get_purchase_records(records) if r['id'] not in completed_ids]

    # If coming from "赎回" button, pre-select the purchase record
    selected_purchase_id = request.args.get('purchase_id', '')
    if selected_purchase_id and selected_purchase_id not in [r['id'] for r in purchase_records]:
        selected_purchase_id = ''  # Reset if completed or invalid

    return render_template('add.html', purchase_records=purchase_records,
                           selected_purchase_id=selected_purchase_id)


@app.route('/edit/<id>', methods=['GET', 'POST'])
def edit_record(id):
    records = load_data()
    record = find_record(records, str(id))
    if not record:
        flash('记录未找到', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            record['product_name'] = request.form['product_name']
            record['amount'] = float(request.form['amount'])
            # annual_rate is entered as percentage, convert to decimal
            raw_rate = request.form['annual_rate'].strip()
            record['annual_rate'] = float(raw_rate) / 100.0 if raw_rate else record['annual_rate']
            record['duration'] = int(request.form['duration'])
            record['purchase_date'] = request.form['purchase_date']
            from utils.helpers import calculate_end_date
            record['end_date'] = calculate_end_date(
                request.form['purchase_date'], int(request.form['duration'])
            )
            record['bank_name'] = request.form.get('bank_name', '未知银行')
            # If it's a redeem record, update profit too
            if record.get('type') == 'redeem':
                if request.form.get('actual_profit'):
                    record['actual_profit'] = float(request.form['actual_profit'])
                record['profit_calc'] = request.form.get('profit_calc', record.get('profit_calc', 'auto'))

            save_data(records)
            flash('记录更新成功！', 'success')
            return redirect(url_for('index'))

        except Exception as e:
            flash(f'更新记录失败: {str(e)}', 'error')
            return redirect(url_for('edit_record', id=id))

    return render_template('edit.html', record=record)


@app.route('/delete/<id>')
def delete_record(id):
    records = load_data()
    record = find_record(records, str(id))
    if not record:
        flash('记录未找到', 'error')
        return redirect(url_for('index'))

    # Delete protection: check for redemptions
    if record.get('type') == 'purchase':
        redeems = get_redeems_for_purchase(records, str(id))
        if redeems:
            flash(f'该产品有 {len(redeems)} 条赎回记录，请先删除赎回记录后再删除该购买记录', 'error')
            return redirect(url_for('records_list'))

    records = [r for r in records if r['id'] != str(id)]
    save_data(records)
    flash('记录删除成功！', 'success')
    return redirect(url_for('index'))


@app.route('/statistics')
def statistics():
    records = load_data()
    today = datetime.now()
    current_year = today.year

    # Time range filter
    range_type = request.args.get('range', 'all')
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    # Filter records by date range
    filtered = records
    if range_type == 'year':
        filtered = [r for r in records if (r.get('purchase_date', '')[:4] == str(current_year)
                    or r.get('redeem_date', '')[:4] == str(current_year))]
    elif range_type == 'month':
        ym = today.strftime('%Y-%m')
        filtered = [r for r in records if (r.get('purchase_date', '')[:7] == ym
                    or r.get('redeem_date', '')[:7] == ym)]
    elif range_type == 'custom' and from_date and to_date:
        filtered = [r for r in records if (
            (r.get('purchase_date', '') >= from_date and r.get('purchase_date', '') <= to_date)
            or (r.get('redeem_date', '') >= from_date and r.get('redeem_date', '') <= to_date)
        )]

    # Charts
    monthly_labels, monthly_profits = calculate_monthly_profits(filtered)
    yearly_labels, yearly_profits = calculate_yearly_profits(filtered)
    product_names, product_amounts = calculate_product_distribution(filtered)

    # Statistics summary
    purchases = get_purchase_records(filtered)
    redeems = [r for r in filtered if r.get('type') == 'redeem']

    total_count = len(purchases)
    redeem_count = len(redeems)
    total_invested = sum(r['amount'] for r in purchases)
    total_profit = sum(r.get('actual_profit', 0) for r in redeems)

    days_list = []
    rates = []
    for r in redeems:
        d = days_between(r['purchase_date'], r['redeem_date'])
        if d > 0:
            days_list.append(d)
            rates.append(calculate_real_rate(r['actual_profit'], r['redeem_amount'], d))

    avg_days = round(sum(days_list) / len(days_list)) if days_list else 0
    avg_rate = round(sum(rates) / len(rates), 2) if rates else 0
    max_rate = round(max(rates), 2) if rates else 0
    profits = [r.get('actual_profit', 0) for r in redeems]
    max_profit = round(max(profits), 2) if profits else 0
    min_profit = round(min(profits), 2) if profits else 0

    # Product ranking (by total profit)
    product_ranking = defaultdict(lambda: {'profit': 0, 'rate': 0, 'days': 0, 'count': 0})
    for r in redeems:
        name = r['product_name']
        product_ranking[name]['profit'] += r.get('actual_profit', 0)
        product_ranking[name]['count'] += 1
        days = days_between(r['purchase_date'], r['redeem_date'])
        if days > 0:
            rate = calculate_real_rate(r['actual_profit'], r['redeem_amount'], days)
            product_ranking[name]['rate'] = rate
            product_ranking[name]['days'] = days

    ranking_by_profit = sorted(product_ranking.items(), key=lambda x: x[1]['profit'], reverse=True)
    ranking_by_rate = sorted(product_ranking.items(), key=lambda x: x[1]['rate'], reverse=True)

    # Bank distribution (from purchases)
    bank_amounts = defaultdict(float)
    for p in purchases:
        bank = p.get('bank_name', '未知银行')
        bank_amounts[bank] += p['amount']
    bank_data = sorted(bank_amounts.items(), key=lambda x: x[1], reverse=True)

    return render_template('statistics.html',
                           range_type=range_type,
                           from_date=from_date,
                           to_date=to_date,
                           monthly_labels=monthly_labels,
                           monthly_profits=monthly_profits,
                           yearly_labels=yearly_labels,
                           yearly_profits=yearly_profits,
                           product_names=product_names,
                           product_amounts=product_amounts,
                           summary=dict(
                               total_count=total_count,
                               redeem_count=redeem_count,
                               total_invested=round(total_invested, 2),
                               total_profit=round(total_profit, 2),
                               avg_days=avg_days,
                               avg_rate=avg_rate,
                               max_rate=max_rate,
                               max_profit=max_profit,
                               min_profit=min_profit,
                           ),
                           ranking_by_profit=ranking_by_profit,
                           ranking_by_rate=ranking_by_rate,
                           bank_data=bank_data)


@app.route('/export')
def export_records():
    fmt = request.args.get('format', 'excel')
    records = load_data()

    # CSV export
    if fmt == 'csv':
        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['类型', '产品名称', '银行', '金额', '年化收益率', '期限(天)',
                         '购买日期', '到期日', '赎回日期', '实际收益', '真实年化', '计算方式'])
        for record in records:
            is_purchase = record.get('type') == 'purchase'
            row = [
                '买入' if is_purchase else '赎回',
                record.get('product_name', ''),
                record.get('bank_name', ''),
                record['amount'] if is_purchase else record.get('redeem_amount', 0),
                f"{record['annual_rate'] * 100:.2f}%" if is_purchase else '',
                record['duration'] if is_purchase else '',
                record['purchase_date'] if is_purchase else '',
                record['end_date'] if is_purchase else '',
                record.get('redeem_date', '') if not is_purchase else '',
                record.get('actual_profit', 0) if not is_purchase else '',
                '',
                record.get('profit_calc', '') if not is_purchase else '',
            ]
            writer.writerow(row)
        csv_data = output.getvalue()
        output.close()
        response = app.response_class(
            csv_data.encode('utf-8-sig'),
            mimetype='text/csv; charset=utf-8-sig',
            headers={'Content-Disposition': 'attachment; filename=finance_records.csv'}
        )
        return response

    # Excel export
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '理财记录'

        headers = ['类型', '产品名称', '银行', '金额', '年化收益率', '期限(天)',
                   '购买日期', '到期日', '赎回日期', '实际收益', '真实年化', '计算方式']
        header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row, record in enumerate(records, 2):
            is_purchase = record.get('type') == 'purchase'
            ws.cell(row=row, column=1, value='买入' if is_purchase else '赎回').border = thin_border
            ws.cell(row=row, column=2, value=record.get('product_name', '')).border = thin_border
            ws.cell(row=row, column=3, value=record.get('bank_name', '')).border = thin_border
            if is_purchase:
                ws.cell(row=row, column=4, value=record['amount']).border = thin_border
                ws.cell(row=row, column=5, value=f"{record['annual_rate'] * 100:.2f}%").border = thin_border
                ws.cell(row=row, column=6, value=record['duration']).border = thin_border
                ws.cell(row=row, column=7, value=record['purchase_date']).border = thin_border
                ws.cell(row=row, column=8, value=record['end_date']).border = thin_border
            else:
                ws.cell(row=row, column=4, value=record.get('redeem_amount', 0)).border = thin_border
                ws.cell(row=row, column=5, value=f"{record.get('annual_rate', 0) * 100:.2f}%").border = thin_border
                ws.cell(row=row, column=6, value=record.get('duration', '')).border = thin_border
                ws.cell(row=row, column=7, value=record.get('purchase_date', '')).border = thin_border
                ws.cell(row=row, column=8, value='').border = thin_border
                ws.cell(row=row, column=9, value=record.get('redeem_date', '')).border = thin_border
                ws.cell(row=row, column=10, value=record.get('actual_profit', 0)).border = thin_border
                ws.cell(row=row, column=11, value='').border = thin_border
                ws.cell(row=row, column=12, value=record.get('profit_calc', 'auto')).border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 4

        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()

        with open(tmp.name, 'rb') as f:
            data = f.read()
        os.unlink(tmp.name)

        return app.response_class(
            data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=finance_records.xlsx'}
        )
    except ImportError:
        flash('Excel 导出需要 openpyxl 库，请运行: pip install openpyxl', 'error')
        return redirect(url_for('records_list'))


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------
@app.errorhandler(404)
def not_found(error):
    return render_template('base.html', error='页面未找到'), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template('base.html', error='服务器内部错误'), 500


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    from werkzeug.serving import is_running_from_reloader
    import signal
    import sys

    def shutdown_handler(signum, frame):
        print("\nServer is shutting down...")
        sys.exit(0)

    if not is_running_from_reloader():
        signal.signal(signal.SIGINT, shutdown_handler)
        signal.signal(signal.SIGTERM, shutdown_handler)

    try:
        app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)
    except Exception as e:
        print(f"Server error: {str(e)}")
    finally:
        print("Server stopped")
